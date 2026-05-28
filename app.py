import io
import os
import secrets
import sqlite3
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from functools import wraps
from pathlib import Path
from zoneinfo import ZoneInfo

from flask import (
    Flask,
    abort,
    flash,
    g,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps, UnidentifiedImageError
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass


Image.MAX_IMAGE_PIXELS = 30_000_000
MONEY_QUANT = Decimal("0.01")
BOGOTA_TZ = ZoneInfo("America/Bogota")

app = Flask(__name__, instance_relative_config=True)
app.config.from_mapping(
    SECRET_KEY=os.environ.get("SECRET_KEY", "dev-change-me-before-deploy"),
    DATABASE=os.environ.get("DATABASE_PATH", os.path.join(app.instance_path, "app.db")),
    UPLOAD_FOLDER=os.environ.get(
        "UPLOAD_FOLDER", os.path.join(app.instance_path, "uploads")
    ),
    MAX_CONTENT_LENGTH=16 * 1024 * 1024,
    IMAGE_MAX_BYTES=500 * 1024,
    INVOICE_MAX_BYTES=10 * 1024 * 1024,
)

STATUS_LABELS = {
    "pending": "待审批",
    "approved": "已同意",
    "rejected": "未同意",
}

STATUS_TRANSLATION_KEYS = {
    "pending": "status_pending",
    "approved": "status_approved",
    "rejected": "status_rejected",
}

STATUS_CLASSES = {
    "pending": "status-pending",
    "approved": "status-approved",
    "rejected": "status-rejected",
}

ROLE_LABELS = {
    "admin": "管理员",
    "approver": "审批人",
    "requester": "申请人",
}

ROLE_TRANSLATION_KEYS = {
    "admin": "role_admin",
    "approver": "role_approver",
    "requester": "role_requester",
}

LANGUAGES = {
    "zh": "中文",
    "es": "Español",
}

TRANSLATIONS = {
    "zh": {
        "app_name": "采购审批",
        "nav_requests": "申请列表",
        "nav_new_request": "新申请",
        "nav_users": "用户管理",
        "logout": "退出",
        "role_admin": "管理员",
        "role_approver": "审批人",
        "role_requester": "申请人",
        "status_pending": "待审批",
        "status_approved": "已同意",
        "status_rejected": "未同意",
        "login": "登录",
        "language": "语言",
        "username": "用户名",
        "password": "密码",
        "enter_system": "进入系统",
        "request_list": "申请列表",
        "request_list_subtitle": "全部采购申请与审批状态",
        "start_date": "开始日期",
        "end_date": "结束日期",
        "status": "状态",
        "all": "全部",
        "filter": "筛选",
        "reset": "重置",
        "export_excel": "导出 Excel",
        "date": "日期",
        "content": "内容",
        "requester": "提交人",
        "item_count": "明细",
        "total_price": "总价",
        "attachments_count": "附件",
        "view": "查看",
        "no_requests": "暂无采购申请",
        "new_request_title": "新申请",
        "new_request_subtitle": "提交采购日期、内容、明细、图片与发票",
        "request_date": "申请日期",
        "purchase_content": "采购内容",
        "purchase_items": "采购明细",
        "add_row": "增加一行",
        "item": "项目",
        "unit_price": "单价",
        "quantity": "数量",
        "line_total": "总价",
        "total": "合计",
        "purchase_item_placeholder": "采购项目",
        "remove": "删除",
        "purchase_images": "采购图片",
        "invoice": "发票",
        "cancel": "取消",
        "submit_request": "提交申请",
        "request_number": "申请",
        "request_total": "申请总价",
        "submitted_at": "提交时间",
        "approver": "审批人",
        "decision_time": "审批时间",
        "decision_comment": "审批备注",
        "no_items": "暂无采购明细",
        "attachments": "附件",
        "pdf": "PDF",
        "image": "图片",
        "no_attachments": "暂无附件",
        "approval": "审批",
        "reject": "未同意",
        "approve": "同意",
        "user_management": "用户管理",
        "user_management_subtitle": "新增用户、调整角色、修改密码",
        "add_user": "新增用户",
        "display_name": "名称",
        "role": "角色",
        "account_status": "状态",
        "created_at": "创建时间",
        "active": "启用",
        "inactive": "停用",
        "edit": "编辑",
        "new_user": "新增用户",
        "edit_user": "编辑用户",
        "user_form_subtitle": "设置登录信息与权限角色",
        "new_password": "新密码",
        "enable_account": "启用账号",
        "save": "保存",
        "invalid_date": "请选择有效的申请日期。",
        "image_unrecognized": "图片文件无法识别，请上传 jpg、png 或 webp。",
        "image_too_large": "图片压缩后仍超过 500KB，请换一张更小的图片。",
        "pdf_too_large": "PDF 发票不能超过 10MB。",
        "file_too_large": "上传内容太大，请减少文件数量或压缩后再上传。",
        "login_failed": "用户名或密码不正确。",
        "request_content_required": "请填写采购内容。",
        "item_description_required": "请填写每一行的采购项目。",
        "items_required": "请至少填写一行采购明细。",
        "number_required": "{field}必须是有效数字。",
        "number_positive": "{field}必须大于 0。",
        "request_created": "采购申请已提交。",
        "request_approved": "已同意该采购申请。",
        "request_rejected": "已标记为未同意。",
        "user_required": "用户名和密码不能为空。",
        "username_required": "用户名不能为空。",
        "self_admin_required": "不能取消当前登录管理员自己的管理员权限或启用状态。",
        "user_created": "用户已新增。",
        "user_updated": "用户信息已更新。",
        "username_exists": "用户名已存在。",
        "export_sheet": "采购申请",
        "export_request_id": "申请编号",
        "export_line_total": "行总价",
        "export_request_total": "申请总价",
    },
    "es": {
        "app_name": "Aprobación de compras",
        "nav_requests": "Solicitudes",
        "nav_new_request": "Nueva solicitud",
        "nav_users": "Usuarios",
        "logout": "Salir",
        "role_admin": "Administrador",
        "role_approver": "Aprobador",
        "role_requester": "Solicitante",
        "status_pending": "Pendiente",
        "status_approved": "Aprobada",
        "status_rejected": "Rechazada",
        "login": "Iniciar sesión",
        "language": "Idioma",
        "username": "Usuario",
        "password": "Contraseña",
        "enter_system": "Entrar",
        "request_list": "Solicitudes",
        "request_list_subtitle": "Todas las solicitudes y su estado",
        "start_date": "Fecha inicial",
        "end_date": "Fecha final",
        "status": "Estado",
        "all": "Todas",
        "filter": "Filtrar",
        "reset": "Restablecer",
        "export_excel": "Exportar Excel",
        "date": "Fecha",
        "content": "Contenido",
        "requester": "Solicitante",
        "item_count": "Líneas",
        "total_price": "Total",
        "attachments_count": "Adjuntos",
        "view": "Ver",
        "no_requests": "No hay solicitudes",
        "new_request_title": "Nueva solicitud",
        "new_request_subtitle": "Ingrese fecha, contenido, detalles, imágenes y factura",
        "request_date": "Fecha de solicitud",
        "purchase_content": "Contenido de compra",
        "purchase_items": "Detalles de compra",
        "add_row": "Agregar línea",
        "item": "Artículo",
        "unit_price": "Precio unitario",
        "quantity": "Cantidad",
        "line_total": "Total",
        "total": "Total",
        "purchase_item_placeholder": "Artículo de compra",
        "remove": "Eliminar",
        "purchase_images": "Imágenes de compra",
        "invoice": "Factura",
        "cancel": "Cancelar",
        "submit_request": "Enviar solicitud",
        "request_number": "Solicitud",
        "request_total": "Total de la solicitud",
        "submitted_at": "Fecha de envío",
        "approver": "Aprobador",
        "decision_time": "Fecha de decisión",
        "decision_comment": "Comentario de aprobación",
        "no_items": "No hay detalles",
        "attachments": "Adjuntos",
        "pdf": "PDF",
        "image": "Imagen",
        "no_attachments": "No hay adjuntos",
        "approval": "Aprobación",
        "reject": "Rechazar",
        "approve": "Aprobar",
        "user_management": "Usuarios",
        "user_management_subtitle": "Crear usuarios, ajustar roles y cambiar contraseñas",
        "add_user": "Agregar usuario",
        "display_name": "Nombre",
        "role": "Rol",
        "account_status": "Estado",
        "created_at": "Creado",
        "active": "Activo",
        "inactive": "Inactivo",
        "edit": "Editar",
        "new_user": "Agregar usuario",
        "edit_user": "Editar usuario",
        "user_form_subtitle": "Configure acceso y permisos",
        "new_password": "Nueva contraseña",
        "enable_account": "Activar cuenta",
        "save": "Guardar",
        "invalid_date": "Seleccione una fecha válida.",
        "image_unrecognized": "No se pudo leer la imagen. Suba jpg, png o webp.",
        "image_too_large": "La imagen aún supera 500KB después de comprimirla.",
        "pdf_too_large": "La factura PDF no puede superar 10MB.",
        "file_too_large": "La carga es demasiado grande. Reduzca archivos o comprímalos.",
        "login_failed": "Usuario o contraseña incorrectos.",
        "request_content_required": "Ingrese el contenido de compra.",
        "item_description_required": "Ingrese el artículo en cada línea.",
        "items_required": "Ingrese al menos una línea de compra.",
        "number_required": "{field} debe ser un número válido.",
        "number_positive": "{field} debe ser mayor que 0.",
        "request_created": "Solicitud enviada.",
        "request_approved": "Solicitud aprobada.",
        "request_rejected": "Solicitud rechazada.",
        "user_required": "Usuario y contraseña son obligatorios.",
        "username_required": "El usuario es obligatorio.",
        "self_admin_required": "No puede quitarse su propio rol de administrador ni desactivar su cuenta.",
        "user_created": "Usuario creado.",
        "user_updated": "Usuario actualizado.",
        "username_exists": "El usuario ya existe.",
        "export_sheet": "Solicitudes",
        "export_request_id": "ID",
        "export_line_total": "Total línea",
        "export_request_total": "Total solicitud",
    },
}

FIXED_USERS = [
    {
        "username": "jose",
        "display_name": "Jose",
        "role": "approver",
        "password": "jose1234",
    },
    {
        "username": "admin",
        "display_name": "Admin",
        "role": "admin",
        "password": "admin1234",
    },
    {
        "username": "carlos",
        "display_name": "Carlos",
        "role": "requester",
        "password": "carlos1234",
    },
]

DEPRECATED_BOOTSTRAP_USERS = ("approver", "user1", "user2")

USERS_TABLE_SQL = """
CREATE TABLE users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    username TEXT NOT NULL UNIQUE,
    display_name TEXT NOT NULL,
    role TEXT NOT NULL CHECK (role IN ('admin', 'approver', 'requester')),
    password_hash TEXT NOT NULL,
    is_active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);
"""

CORE_SCHEMA = """
CREATE TABLE IF NOT EXISTS purchase_requests (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    requester_id INTEGER NOT NULL REFERENCES users(id),
    request_date TEXT NOT NULL,
    content TEXT NOT NULL,
    status TEXT NOT NULL DEFAULT 'pending'
        CHECK (status IN ('pending', 'approved', 'rejected')),
    decision_comment TEXT,
    decided_by INTEGER REFERENCES users(id),
    decided_at TEXT,
    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS purchase_request_items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    request_id INTEGER NOT NULL REFERENCES purchase_requests(id)
        ON DELETE CASCADE,
    description TEXT NOT NULL,
    unit_price_cents INTEGER NOT NULL,
    quantity TEXT NOT NULL,
    line_total_cents INTEGER NOT NULL,
    sort_order INTEGER NOT NULL DEFAULT 0
);

CREATE TABLE IF NOT EXISTS attachments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    request_id INTEGER NOT NULL REFERENCES purchase_requests(id)
        ON DELETE CASCADE,
    kind TEXT NOT NULL CHECK (kind IN ('image', 'invoice')),
    original_name TEXT NOT NULL,
    stored_name TEXT NOT NULL UNIQUE,
    mime_type TEXT NOT NULL,
    size_bytes INTEGER NOT NULL,
    uploaded_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_purchase_requests_status
    ON purchase_requests(status);
CREATE INDEX IF NOT EXISTS idx_purchase_requests_request_date
    ON purchase_requests(request_date);
CREATE INDEX IF NOT EXISTS idx_purchase_requests_created_at
    ON purchase_requests(created_at);
CREATE INDEX IF NOT EXISTS idx_items_request_id
    ON purchase_request_items(request_id);
CREATE INDEX IF NOT EXISTS idx_attachments_request_id
    ON attachments(request_id);
"""


class ValidationError(Exception):
    pass


def current_language() -> str:
    try:
        return g.get("lang", "zh")
    except RuntimeError:
        return "zh"


def t(key: str, **kwargs) -> str:
    lang = current_language()
    text = TRANSLATIONS.get(lang, TRANSLATIONS["zh"]).get(
        key, TRANSLATIONS["zh"].get(key, key)
    )
    if kwargs:
        return text.format(**kwargs)
    return text


def status_label(status: str) -> str:
    return t(STATUS_TRANSLATION_KEYS.get(status, status))


def role_label(role: str) -> str:
    return t(ROLE_TRANSLATION_KEYS.get(role, role))


def localized_status_labels() -> dict[str, str]:
    return {status: status_label(status) for status in STATUS_LABELS}


def localized_role_labels() -> dict[str, str]:
    return {role: role_label(role) for role in ROLE_LABELS}


def local_today() -> date:
    return datetime.now(BOGOTA_TZ).date()


def utc_now() -> str:
    return datetime.now(BOGOTA_TZ).replace(microsecond=0).isoformat()


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(app.config["DATABASE"])
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(error=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def table_exists(db, table_name: str) -> bool:
    return (
        db.execute(
            "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
            (table_name,),
        ).fetchone()
        is not None
    )


def table_columns(db, table_name: str) -> set[str]:
    if not table_exists(db, table_name):
        return set()
    return {row["name"] for row in db.execute(f"PRAGMA table_info({table_name})")}


def users_table_supports_roles(db) -> bool:
    row = db.execute(
        "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'users'"
    ).fetchone()
    if row is None or row["sql"] is None:
        return False
    sql = row["sql"].lower()
    return "admin" in sql and "requester" in sql and "is_active" in table_columns(db, "users")


def normalize_role(role: str) -> str:
    if role == "user":
        return "requester"
    if role in ROLE_LABELS:
        return role
    return "requester"


def ensure_users_schema(db):
    if not table_exists(db, "users"):
        db.execute(USERS_TABLE_SQL)
        return

    if users_table_supports_roles(db):
        return

    existing_columns = table_columns(db, "users")
    rows = db.execute("SELECT * FROM users ORDER BY id").fetchall()

    db.execute("CREATE TABLE users_new " + USERS_TABLE_SQL.split("users", 1)[1])
    for row in rows:
        db.execute(
            """
            INSERT INTO users_new
                (id, username, display_name, role, password_hash, is_active, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row["id"],
                row["username"],
                row["display_name"],
                normalize_role(row["role"]),
                row["password_hash"],
                row["is_active"] if "is_active" in existing_columns else 1,
                row["created_at"]
                if "created_at" in existing_columns and row["created_at"]
                else utc_now(),
            ),
        )

    db.execute("DROP TABLE users")
    db.execute("ALTER TABLE users_new RENAME TO users")


def seed_users(db):
    for user in FIXED_USERS:
        existing = db.execute(
            "SELECT id FROM users WHERE username = ?", (user["username"],)
        ).fetchone()

        if existing:
            db.execute(
                """
                UPDATE users
                SET display_name = ?, role = ?, is_active = 1
                WHERE username = ?
                """,
                (user["display_name"], user["role"], user["username"]),
            )
        else:
            db.execute(
                """
                INSERT INTO users
                    (username, display_name, role, password_hash, is_active)
                VALUES (?, ?, ?, ?, 1)
                """,
                (
                    user["username"],
                    user["display_name"],
                    user["role"],
                    generate_password_hash(user["password"]),
                ),
            )

    db.execute(
        f"""
        UPDATE users
        SET is_active = 0
        WHERE username IN ({",".join("?" for _ in DEPRECATED_BOOTSTRAP_USERS)})
        """,
        DEPRECATED_BOOTSTRAP_USERS,
    )


def init_db():
    os.makedirs(app.instance_path, exist_ok=True)
    Path(app.config["DATABASE"]).parent.mkdir(parents=True, exist_ok=True)
    Path(app.config["UPLOAD_FOLDER"]).mkdir(parents=True, exist_ok=True)

    db = sqlite3.connect(app.config["DATABASE"])
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys = OFF")
    ensure_users_schema(db)
    db.executescript(CORE_SCHEMA)
    seed_users(db)
    db.commit()
    db.execute("PRAGMA foreign_keys = ON")
    db.close()


@app.before_request
def load_current_user():
    requested_lang = request.values.get("lang")
    if requested_lang in LANGUAGES:
        session["lang"] = requested_lang
    g.lang = session.get("lang", "zh")
    if g.lang not in LANGUAGES:
        g.lang = "zh"
        session["lang"] = "zh"

    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_urlsafe(32)

    user_id = session.get("user_id")
    g.user = None
    if user_id is not None:
        g.user = get_db().execute(
            """
            SELECT id, username, display_name, role, is_active
            FROM users
            WHERE id = ? AND is_active = 1
            """,
            (user_id,),
        ).fetchone()
        if g.user is None:
            session.clear()


def can_approve(user=None) -> bool:
    user = user or g.get("user")
    return bool(user and user["role"] in {"admin", "approver"})


def can_manage_users(user=None) -> bool:
    user = user or g.get("user")
    return bool(user and user["role"] == "admin")


def can_submit_requests(user=None) -> bool:
    user = user or g.get("user")
    return bool(user and user["role"] == "requester")


@app.context_processor
def inject_globals():
    user = g.get("user")
    return {
        "current_user": user,
        "csrf_token": session.get("csrf_token"),
        "lang": current_language(),
        "languages": LANGUAGES,
        "t": t,
        "status_labels": localized_status_labels(),
        "status_classes": STATUS_CLASSES,
        "role_labels": localized_role_labels(),
        "can_approve_requests": can_approve(user),
        "can_manage_user_accounts": can_manage_users(user),
        "can_submit_purchase_requests": can_submit_requests(user),
    }


def login_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if g.user is None:
            return redirect(url_for("login", next=request.path))
        return view(**kwargs)

    return wrapped_view


def approver_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if g.user is None:
            return redirect(url_for("login", next=request.path))
        if not can_approve(g.user):
            abort(403)
        return view(**kwargs)

    return wrapped_view


def admin_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if g.user is None:
            return redirect(url_for("login", next=request.path))
        if not can_manage_users(g.user):
            abort(403)
        return view(**kwargs)

    return wrapped_view


def requester_required(view):
    @wraps(view)
    def wrapped_view(**kwargs):
        if g.user is None:
            return redirect(url_for("login", next=request.path))
        if not can_submit_requests(g.user):
            abort(403)
        return view(**kwargs)

    return wrapped_view


def validate_csrf():
    token = request.form.get("csrf_token")
    if not token or token != session.get("csrf_token"):
        abort(400)


def request_summary(content: str) -> str:
    cleaned = " ".join(content.split())
    if len(cleaned) <= 42:
        return cleaned
    return cleaned[:42] + "..."


def parse_request_date(value: str) -> str:
    try:
        return date.fromisoformat(value).isoformat()
    except (TypeError, ValueError):
        raise ValidationError(t("invalid_date"))


def parse_optional_date(value: str) -> str | None:
    value = (value or "").strip()
    if not value:
        return None
    try:
        return date.fromisoformat(value).isoformat()
    except ValueError:
        return None


def cents_to_decimal(cents: int | None) -> Decimal:
    return (Decimal(cents or 0) / Decimal(100)).quantize(MONEY_QUANT)


def decimal_to_cents(value: Decimal) -> int:
    return int((value * Decimal(100)).quantize(Decimal("1"), ROUND_HALF_UP))


def parse_positive_decimal(value: str, field_name: str) -> Decimal:
    try:
        parsed = Decimal((value or "").strip().replace(",", ""))
    except InvalidOperation:
        raise ValidationError(t("number_required", field=field_name))

    if parsed <= 0:
        raise ValidationError(t("number_positive", field=field_name))
    return parsed


def decimal_text(value: Decimal) -> str:
    text = format(value.normalize(), "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def extract_form_items(form) -> list[dict[str, str]]:
    descriptions = form.getlist("item_description")
    unit_prices = form.getlist("unit_price")
    quantities = form.getlist("quantity")
    row_count = max(1, len(descriptions), len(unit_prices), len(quantities))
    items = []

    for index in range(row_count):
        items.append(
            {
                "description": descriptions[index].strip()
                if index < len(descriptions)
                else "",
                "unit_price": unit_prices[index].strip()
                if index < len(unit_prices)
                else "",
                "quantity": quantities[index].strip() if index < len(quantities) else "",
            }
        )
    return items


def parse_items(posted_items: list[dict[str, str]]) -> list[dict[str, object]]:
    parsed_items = []
    for item in posted_items:
        if not item["description"] and not item["unit_price"] and not item["quantity"]:
            continue

        if not item["description"]:
            raise ValidationError(t("item_description_required"))

        unit_price = parse_positive_decimal(item["unit_price"], t("unit_price"))
        quantity = parse_positive_decimal(item["quantity"], t("quantity"))
        line_total = unit_price * quantity

        parsed_items.append(
            {
                "description": item["description"],
                "unit_price_cents": decimal_to_cents(unit_price),
                "quantity": decimal_text(quantity),
                "line_total_cents": decimal_to_cents(line_total),
            }
        )

    if not parsed_items:
        raise ValidationError(t("items_required"))
    return parsed_items


def build_filters(args):
    selected_status = args.get("status", "all")
    if selected_status not in {"all", *STATUS_LABELS.keys()}:
        selected_status = "all"

    return {
        "status": selected_status,
        "date_from": parse_optional_date(args.get("date_from", "")),
        "date_to": parse_optional_date(args.get("date_to", "")),
    }


def filter_params(filters, include_status=True):
    clauses = []
    params = []
    if include_status and filters["status"] != "all":
        clauses.append("pr.status = ?")
        params.append(filters["status"])
    if filters["date_from"]:
        clauses.append("pr.request_date >= ?")
        params.append(filters["date_from"])
    if filters["date_to"]:
        clauses.append("pr.request_date <= ?")
        params.append(filters["date_to"])
    return clauses, params


def status_counts(filters):
    clauses, params = filter_params(filters, include_status=False)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    rows = get_db().execute(
        f"""
        SELECT pr.status, COUNT(*) AS count
        FROM purchase_requests pr
        {where}
        GROUP BY pr.status
        """,
        params,
    ).fetchall()
    counts = {status: 0 for status in STATUS_LABELS}
    for row in rows:
        counts[row["status"]] = row["count"]
    counts["all"] = sum(counts.values())
    return counts


def normalize_image(image: Image.Image) -> Image.Image:
    image = ImageOps.exif_transpose(image)
    image.thumbnail((2400, 2400), Image.Resampling.LANCZOS)

    if image.mode in ("RGBA", "LA") or (
        image.mode == "P" and "transparency" in image.info
    ):
        background = Image.new("RGB", image.size, (255, 255, 255))
        alpha = image.convert("RGBA").getchannel("A")
        background.paste(image.convert("RGBA"), mask=alpha)
        return background

    return image.convert("RGB")


def compress_image(file_storage) -> tuple[str, str, int]:
    try:
        file_storage.stream.seek(0)
        with Image.open(file_storage.stream) as opened:
            original = normalize_image(opened)
    except (UnidentifiedImageError, OSError):
        raise ValidationError(t("image_unrecognized"))

    working = original.copy()
    quality = 86
    scale = 1.0
    payload = None

    for _ in range(36):
        buffer = io.BytesIO()
        working.save(buffer, "JPEG", quality=quality, optimize=True, progressive=True)
        size = buffer.tell()

        if size <= app.config["IMAGE_MAX_BYTES"]:
            payload = buffer.getvalue()
            break

        if quality > 42:
            quality -= 8
            continue

        scale *= 0.82
        width = max(240, int(original.width * scale))
        height = max(240, int(original.height * scale))
        if width == working.width and height == working.height:
            quality = max(30, quality - 4)
            continue
        working = original.resize((width, height), Image.Resampling.LANCZOS)
        quality = 78

    if payload is None:
        raise ValidationError(t("image_too_large"))

    stored_name = f"{uuid.uuid4().hex}.jpg"
    path = os.path.join(app.config["UPLOAD_FOLDER"], stored_name)
    with open(path, "wb") as output:
        output.write(payload)

    return stored_name, "image/jpeg", len(payload)


def save_invoice_pdf(file_storage) -> tuple[str, str, int]:
    file_storage.stream.seek(0)
    payload = file_storage.read()
    if len(payload) > app.config["INVOICE_MAX_BYTES"]:
        raise ValidationError(t("pdf_too_large"))

    stored_name = f"{uuid.uuid4().hex}.pdf"
    path = os.path.join(app.config["UPLOAD_FOLDER"], stored_name)
    with open(path, "wb") as output:
        output.write(payload)

    return stored_name, "application/pdf", len(payload)


def save_attachment(file_storage, kind: str):
    original_name = secure_filename(file_storage.filename or "")
    if not original_name:
        return None

    suffix = Path(original_name).suffix.lower()
    if kind == "invoice" and suffix == ".pdf":
        stored_name, mime_type, size_bytes = save_invoice_pdf(file_storage)
    else:
        stored_name, mime_type, size_bytes = compress_image(file_storage)

    return {
        "kind": kind,
        "original_name": original_name,
        "stored_name": stored_name,
        "mime_type": mime_type,
        "size_bytes": size_bytes,
    }


def insert_attachment(db, request_id: int, attachment):
    if attachment is None:
        return

    db.execute(
        """
        INSERT INTO attachments
            (request_id, kind, original_name, stored_name, mime_type, size_bytes)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            request_id,
            attachment["kind"],
            attachment["original_name"],
            attachment["stored_name"],
            attachment["mime_type"],
            attachment["size_bytes"],
        ),
    )


def insert_items(db, request_id: int, items):
    for index, item in enumerate(items):
        db.execute(
            """
            INSERT INTO purchase_request_items
                (request_id, description, unit_price_cents, quantity,
                 line_total_cents, sort_order)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                request_id,
                item["description"],
                item["unit_price_cents"],
                item["quantity"],
                item["line_total_cents"],
                index,
            ),
        )


def load_request(request_id: int):
    purchase_request = get_db().execute(
        """
        SELECT
            pr.*,
            requester.display_name AS requester_name,
            approver.display_name AS approver_name,
            COALESCE(item_totals.total_cents, 0) AS total_cents
        FROM purchase_requests pr
        JOIN users requester ON requester.id = pr.requester_id
        LEFT JOIN users approver ON approver.id = pr.decided_by
        LEFT JOIN (
            SELECT request_id, SUM(line_total_cents) AS total_cents
            FROM purchase_request_items
            GROUP BY request_id
        ) item_totals ON item_totals.request_id = pr.id
        WHERE pr.id = ?
        """,
        (request_id,),
    ).fetchone()

    if purchase_request is None:
        abort(404)
    return purchase_request


def load_request_items(request_id: int):
    return get_db().execute(
        """
        SELECT *
        FROM purchase_request_items
        WHERE request_id = ?
        ORDER BY sort_order, id
        """,
        (request_id,),
    ).fetchall()


@app.template_filter("bytes")
def format_bytes(value):
    value = int(value or 0)
    if value < 1024:
        return f"{value} B"
    if value < 1024 * 1024:
        return f"{value / 1024:.1f} KB"
    return f"{value / 1024 / 1024:.1f} MB"


@app.template_filter("datetime")
def format_datetime(value):
    if not value:
        return ""
    raw_value = str(value)
    try:
        parsed = datetime.fromisoformat(raw_value.replace("Z", "+00:00"))
        if parsed.tzinfo is not None:
            parsed = parsed.astimezone(BOGOTA_TZ)
        return parsed.strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return raw_value.replace("T", " ").replace("Z", "")


@app.template_filter("money")
def format_money(value):
    return f"{cents_to_decimal(value):,.2f}"


@app.route("/login", methods=("GET", "POST"))
def login():
    if g.user is not None:
        return redirect(url_for("index"))

    if request.method == "POST":
        validate_csrf()
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = get_db().execute(
            "SELECT * FROM users WHERE username = ? AND is_active = 1", (username,)
        ).fetchone()

        if user is None or not check_password_hash(user["password_hash"], password):
            flash(t("login_failed"), "error")
        else:
            selected_lang = current_language()
            session.clear()
            session["lang"] = selected_lang
            session["user_id"] = user["id"]
            session["csrf_token"] = secrets.token_urlsafe(32)
            return redirect(request.args.get("next") or url_for("index"))

    return render_template("login.html")


@app.route("/logout", methods=("POST",))
@login_required
def logout():
    validate_csrf()
    selected_lang = current_language()
    session.clear()
    session["lang"] = selected_lang
    return redirect(url_for("login"))


@app.route("/")
@login_required
def index():
    filters = build_filters(request.args)
    clauses, params = filter_params(filters)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""

    requests = get_db().execute(
        f"""
        SELECT
            pr.*,
            requester.display_name AS requester_name,
            COALESCE(attachments.attachment_count, 0) AS attachment_count,
            COALESCE(items.item_count, 0) AS item_count,
            COALESCE(items.total_cents, 0) AS total_cents
        FROM purchase_requests pr
        JOIN users requester ON requester.id = pr.requester_id
        LEFT JOIN (
            SELECT request_id, COUNT(*) AS attachment_count
            FROM attachments
            GROUP BY request_id
        ) attachments ON attachments.request_id = pr.id
        LEFT JOIN (
            SELECT
                request_id,
                COUNT(*) AS item_count,
                SUM(line_total_cents) AS total_cents
            FROM purchase_request_items
            GROUP BY request_id
        ) items ON items.request_id = pr.id
        {where}
        ORDER BY pr.request_date DESC, pr.created_at DESC, pr.id DESC
        """,
        params,
    ).fetchall()

    return render_template(
        "index.html",
        requests=requests,
        counts=status_counts(filters),
        filters=filters,
        summary=request_summary,
    )


@app.route("/requests/new", methods=("GET", "POST"))
@requester_required
def new_request():
    posted_items = extract_form_items(request.form) if request.method == "POST" else []

    if request.method == "POST":
        validate_csrf()
        content = request.form.get("content", "").strip()
        request_date = request.form.get("request_date", "")

        try:
            parsed_date = parse_request_date(request_date)
            if not content:
                raise ValidationError(t("request_content_required"))
            items = parse_items(posted_items)

            saved_files = []
            db = get_db()
            cursor = db.execute(
                """
                INSERT INTO purchase_requests
                    (requester_id, request_date, content, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (g.user["id"], parsed_date, content, utc_now(), utc_now()),
            )
            request_id = cursor.lastrowid
            insert_items(db, request_id, items)

            for image in request.files.getlist("images"):
                attachment = save_attachment(image, "image")
                if attachment:
                    saved_files.append(attachment["stored_name"])
                    insert_attachment(db, request_id, attachment)

            for invoice in request.files.getlist("invoice"):
                attachment = save_attachment(invoice, "invoice")
                if attachment:
                    saved_files.append(attachment["stored_name"])
                    insert_attachment(db, request_id, attachment)

            db.commit()
            flash(t("request_created"), "success")
            return redirect(url_for("request_detail", request_id=request_id))
        except ValidationError as error:
            get_db().rollback()
            for stored_name in locals().get("saved_files", []):
                path = os.path.join(app.config["UPLOAD_FOLDER"], stored_name)
                if os.path.exists(path):
                    os.remove(path)
            flash(str(error), "error")

    if not posted_items:
        posted_items = [{"description": "", "unit_price": "", "quantity": ""}]

    return render_template(
        "request_form.html",
        today=local_today().isoformat(),
        content=request.form.get("content", ""),
        request_date=request.form.get("request_date", local_today().isoformat()),
        items=posted_items,
    )


@app.route("/requests/<int:request_id>")
@login_required
def request_detail(request_id):
    purchase_request = load_request(request_id)
    items = load_request_items(request_id)
    attachments = get_db().execute(
        "SELECT * FROM attachments WHERE request_id = ? ORDER BY id",
        (request_id,),
    ).fetchall()

    return render_template(
        "request_detail.html",
        purchase_request=purchase_request,
        items=items,
        attachments=attachments,
    )


@app.route("/requests/<int:request_id>/approve", methods=("POST",))
@approver_required
def approve_request(request_id):
    validate_csrf()
    decision_comment = request.form.get("decision_comment", "").strip() or None
    load_request(request_id)
    get_db().execute(
        """
        UPDATE purchase_requests
        SET status = 'approved',
            decision_comment = ?,
            decided_by = ?,
            decided_at = ?,
            updated_at = ?
        WHERE id = ?
        """,
        (decision_comment, g.user["id"], utc_now(), utc_now(), request_id),
    )
    get_db().commit()
    flash(t("request_approved"), "success")
    return redirect(url_for("request_detail", request_id=request_id))


@app.route("/requests/<int:request_id>/reject", methods=("POST",))
@approver_required
def reject_request(request_id):
    validate_csrf()
    decision_comment = request.form.get("decision_comment", "").strip() or None
    load_request(request_id)
    get_db().execute(
        """
        UPDATE purchase_requests
        SET status = 'rejected',
            decision_comment = ?,
            decided_by = ?,
            decided_at = ?,
            updated_at = ?
        WHERE id = ?
        """,
        (decision_comment, g.user["id"], utc_now(), utc_now(), request_id),
    )
    get_db().commit()
    flash(t("request_rejected"), "success")
    return redirect(url_for("request_detail", request_id=request_id))


@app.route("/users")
@admin_required
def users():
    users_list = get_db().execute(
        """
        SELECT id, username, display_name, role, is_active, created_at
        FROM users
        ORDER BY is_active DESC, role, username
        """
    ).fetchall()
    return render_template("users.html", users=users_list)


@app.route("/users/new", methods=("GET", "POST"))
@admin_required
def new_user():
    form = {
        "username": "",
        "display_name": "",
        "role": "requester",
        "password": "",
        "is_active": "1",
    }

    if request.method == "POST":
        validate_csrf()
        form.update(
            {
                "username": request.form.get("username", "").strip(),
                "display_name": request.form.get("display_name", "").strip(),
                "role": request.form.get("role", "requester"),
                "password": request.form.get("password", ""),
                "is_active": "1" if request.form.get("is_active") else "0",
            }
        )

        if form["role"] not in ROLE_LABELS:
            form["role"] = "requester"

        if not form["username"] or not form["password"]:
            flash(t("user_required"), "error")
        else:
            try:
                get_db().execute(
                    """
                    INSERT INTO users
                        (username, display_name, role, password_hash, is_active,
                         created_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        form["username"],
                        form["display_name"] or form["username"],
                        form["role"],
                        generate_password_hash(form["password"]),
                        1 if form["is_active"] == "1" else 0,
                        utc_now(),
                    ),
                )
                get_db().commit()
                flash(t("user_created"), "success")
                return redirect(url_for("users"))
            except sqlite3.IntegrityError:
                flash(t("username_exists"), "error")

    return render_template("user_form.html", form=form, mode="new")


@app.route("/users/<int:user_id>/edit", methods=("GET", "POST"))
@admin_required
def edit_user(user_id):
    user = get_db().execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if user is None:
        abort(404)

    form = {
        "username": user["username"],
        "display_name": user["display_name"],
        "role": user["role"],
        "password": "",
        "is_active": "1" if user["is_active"] else "0",
    }

    if request.method == "POST":
        validate_csrf()
        form.update(
            {
                "username": request.form.get("username", "").strip(),
                "display_name": request.form.get("display_name", "").strip(),
                "role": request.form.get("role", "requester"),
                "password": request.form.get("password", ""),
                "is_active": "1" if request.form.get("is_active") else "0",
            }
        )

        if form["role"] not in ROLE_LABELS:
            form["role"] = "requester"

        if not form["username"]:
            flash(t("username_required"), "error")
        elif user_id == g.user["id"] and (
            form["role"] != "admin" or form["is_active"] != "1"
        ):
            flash(t("self_admin_required"), "error")
        else:
            try:
                get_db().execute(
                    """
                    UPDATE users
                    SET username = ?,
                        display_name = ?,
                        role = ?,
                        is_active = ?
                    WHERE id = ?
                    """,
                    (
                        form["username"],
                        form["display_name"] or form["username"],
                        form["role"],
                        1 if form["is_active"] == "1" else 0,
                        user_id,
                    ),
                )
                if form["password"]:
                    get_db().execute(
                        "UPDATE users SET password_hash = ? WHERE id = ?",
                        (generate_password_hash(form["password"]), user_id),
                    )
                get_db().commit()
                flash(t("user_updated"), "success")
                return redirect(url_for("users"))
            except sqlite3.IntegrityError:
                flash(t("username_exists"), "error")

    return render_template("user_form.html", form=form, mode="edit", edited_user=user)


def export_rows(filters):
    clauses, params = filter_params(filters)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    return get_db().execute(
        f"""
        SELECT
            pr.id,
            pr.request_date,
            pr.content,
            pr.status,
            pr.decision_comment,
            pr.decided_at,
            pr.created_at,
            requester.display_name AS requester_name,
            approver.display_name AS approver_name,
            item.description AS item_description,
            item.unit_price_cents,
            item.quantity,
            item.line_total_cents,
            totals.total_cents
        FROM purchase_requests pr
        JOIN users requester ON requester.id = pr.requester_id
        LEFT JOIN users approver ON approver.id = pr.decided_by
        LEFT JOIN purchase_request_items item ON item.request_id = pr.id
        LEFT JOIN (
            SELECT request_id, SUM(line_total_cents) AS total_cents
            FROM purchase_request_items
            GROUP BY request_id
        ) totals ON totals.request_id = pr.id
        {where}
        ORDER BY pr.request_date DESC, pr.id DESC, item.sort_order, item.id
        """,
        params,
    ).fetchall()


@app.route("/export.xlsx")
@admin_required
def export_requests():
    filters = build_filters(request.args)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = t("export_sheet")

    headers = [
        t("export_request_id"),
        t("request_date"),
        t("requester"),
        t("purchase_content"),
        t("item"),
        t("unit_price"),
        t("quantity"),
        t("export_line_total"),
        t("export_request_total"),
        t("status"),
        t("approver"),
        t("decision_time"),
        t("decision_comment"),
        t("submitted_at"),
    ]
    sheet.append(headers)

    for row in export_rows(filters):
        sheet.append(
            [
                row["id"],
                row["request_date"],
                row["requester_name"],
                row["content"],
                row["item_description"] or "",
                float(cents_to_decimal(row["unit_price_cents"]))
                if row["unit_price_cents"] is not None
                else "",
                float(row["quantity"]) if row["quantity"] else "",
                float(cents_to_decimal(row["line_total_cents"]))
                if row["line_total_cents"] is not None
                else "",
                float(cents_to_decimal(row["total_cents"])),
                status_label(row["status"]),
                row["approver_name"] or "",
                format_datetime(row["decided_at"]),
                row["decision_comment"] or "",
                format_datetime(row["created_at"]),
            ]
        )

    header_fill = PatternFill("solid", fgColor="2F7D57")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = [10, 13, 14, 32, 26, 12, 10, 12, 12, 10, 12, 20, 30, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in sheet.iter_rows(min_row=2):
        row[3].alignment = Alignment(wrap_text=True, vertical="top")
        row[4].alignment = Alignment(wrap_text=True, vertical="top")
        row[12].alignment = Alignment(wrap_text=True, vertical="top")
        for money_cell in (row[5], row[7], row[8]):
            money_cell.number_format = '#,##0.00'
        row[6].number_format = '#,##0.##'

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"procurement-requests-{local_today().isoformat()}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/uploads/<path:filename>")
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


@app.errorhandler(413)
def payload_too_large(error):
    flash(t("file_too_large"), "error")
    return redirect(request.referrer or url_for("new_request"))


init_db()


if __name__ == "__main__":
    app.run(debug=True)
